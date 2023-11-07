import numpy as np
import tkinter as tk
from tkinter import ttk
import tensorflow as tf
from PIL import Image, ImageTk
import matplotlib.pyplot as plt
import mysql.connector
import openpyxl

# Datos de conexión a MySQL
db_host = "localhost"
db_user = "danny"
db_password = "danny"
db_name = "sistema_consumo"

# Crear una conexión a la base de datos
conexion = mysql.connector.connect(
    host=db_host,
    user=db_user,
    password=db_password,
    database=db_name
)

# Crear un cursor para ejecutar consultas SQL
cursor = conexion.cursor()

# Datos de entrada
años = np.array([1985, 1986, 1987, 1988, 1989, 1990, 1991, 1992, 1993, 1994, 1995, 1996, 1997, 1998, 1999, 2000, 2001, 2002, 2003, 2004, 2005, 2006, 2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017])
consumos = np.array([126.56, 128.53, 132.66, 138.75, 147.05, 151.21, 154.75, 156.69, 155.46, 160.90, 165.60, 173.37, 189.14, 193.39, 205.63, 220.92, 233.07, 239.91, 257.30, 273.65, 284.46, 295.26, 301.53, 310.67, 291.53, 298.15, 291.37, 293.77, 281.47, 274.94, 277.48, 271.19, 272.90])

# Crear y entrenar un modelo de regresión con TensorFlow
model_tf = tf.keras.Sequential([
    tf.keras.layers.Dense(units=1, input_shape=(1,))
])
model_tf.compile(optimizer='adam', loss='mean_squared_error')
model_tf.fit(años, consumos, epochs=1000, verbose=0)

# Función para predecir el consumo utilizando TensorFlow
def predecir_consumo_tensorflow():
    año_a_predecir = float(entry_año.get())
    consumo_predicho = model_tf.predict([año_a_predecir])
    label_resultado.config(text=f"Consumo estimado (TensorFlow): {consumo_predicho[0][0]:.2f}")


    # Crear un gráfico de dispersión de los datos reales
    plt.figure(figsize=(8, 4))
    plt.scatter(años, consumos, label="Datos reales", color='b')
    
    # Agregar la predicción al gráfico
    plt.scatter(año_a_predecir, consumo_predicho[0][0], label="Predicción", color='r')
    
    # Configuración del gráfico
    plt.xlabel('Año')
    plt.ylabel('Consumo eléctrico')
    plt.title('Predicción de Consumo Eléctrico')
    plt.legend()
    plt.grid(True)

    # Mostrar el gráfico
    plt.show()


def insertar():
    try:
        año = float(entry_año.get())
        consumo = float(label_resultado.cget("text").split(":")[1].strip())  # Obtiene el valor del label_resultado
         # Ejecuta la consulta para insertar en la tabla de consumo
        consulta = "INSERT INTO consumido (año, cosumo) VALUES (%s, %s)"
        valores = (año, consumo)
        cursor.execute(consulta, valores)
        conexion.commit()

        label_resultado.config(text="Consumo registrado en la base de datos")
        # Intenta ejecutar una consulta simple para verificar la conexión
        #cursor.execute("SELECT 1")
        #label_resultado.config(text="Conexión a la base de datos exitosa")
    except mysql.connector.Error as error:
        label_resultado.config(text=f"Error de conexión a la base de datos: {error}")


def exportar():
    try: 
      cursor.execute("SELECT año, cosumo FROM consumido")
      resultados = cursor.fetchall()
      #label_resultado.config(text="Archivo exportado")
      # Crear un nuevo libro de trabajo de Excel y una hoja de cálculo
      workbook = openpyxl.Workbook()
      sheet = workbook.active

      # Escribir los resultados en la hoja de cálculo
      sheet.cell(row=1, column=1, value="Año")
      sheet.cell(row=1, column=2, value="Consumo")

      for idx, resultado in enumerate(resultados, start=2):
            año, consumo = resultado
            sheet.cell(row=idx, column=1, value=año)
            sheet.cell(row=idx, column=2, value=consumo)

        # Guardar el archivo Excel
      workbook.save("resultados_consumo.xlsx")

      label_resultado.config(text="Resultados exportados a resultados_consumo.xlsx")
    except mysql.connector.Error as error:
        label_resultado.config(text=f"Error al exportar los resultados: {error}")


# Función para mostrar los resultados de la tabla "consumido"
def mostrar_resultados():
    try:
        # Ejecuta una consulta SQL para obtener todos los registros de la tabla "consumido"
        cursor.execute("SELECT año, cosumo FROM consumido")
        resultados = cursor.fetchall()

        # Borra todos los elementos actuales en la tabla
        for i in tabla.get_children():
            tabla.delete(i)

        # Inserta los nuevos resultados en la tabla
        for resultado in resultados:
            tabla.insert("", "end", values=resultado)
            #label_resultado.config(text="Resultados de la consulta:")
    except mysql.connector.Error as error:
        label_resultado.config(text=f"Error al obtener los resultados: {error}")




    
# Crear la ventana de la aplicación
ventana = tk.Tk()
ventana.title("Predicción de Consumo Eléctrico")
ventana.geometry("400x600")

##################################################################
# Crear un Treeview (tabla)
tabla = ttk.Treeview(ventana, columns=("Año", "Consumo"), show="headings")
tabla.heading("#1", text="Año")
tabla.heading("#2", text="Consumo")
tabla.column("#1", width=100)
tabla.column("#2", width=100)
mostrar_resultados()
##################################################################

# Crear elementos de la interfaz
label_año = tk.Label(ventana, text="Año de consumo a predecir:")
entry_año = tk.Entry(ventana)
boton_predecir_tensorflow = tk.Button(ventana, text="Predecir (TensorFlow)", command=predecir_consumo_tensorflow)
boton = tk.Button(ventana, text="Exportar", command=exportar) #boton de exportar
boton_registrar = tk.Button(ventana, text="Registrar consumo", command=insertar) #boton de exportar
label_resultado = tk.Label(ventana, text="")
label_imagen = tk.Label(ventana)

# Cargar una imagen para mostrar
imagen = Image.open("C:/xampp/htdocs/tensorflow/tensorflow/tensorflow/static/imgs/lineas.jpg")
imagen = imagen.resize((400, 200), Image.ANTIALIAS)
imagen = ImageTk.PhotoImage(imagen)
label_imagen.config(image=imagen)
label_imagen.image = imagen

# Colocar elementos en la ventana
label_imagen.pack()
label_año.pack()
entry_año.pack()
boton_predecir_tensorflow.pack()
label_resultado.pack()
tabla.pack()
boton_registrar.pack()
boton.pack()


# Iniciar la ventana
ventana.mainloop()

